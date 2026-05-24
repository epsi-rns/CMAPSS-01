# build_xlsx_v6_pca.py
#
# Adds PCA sheets to cmapss_FDxxx_v5.xlsx → cmapss_FDxxx_v6.xlsx
# for all four C-MAPSS sub-datasets.
#
# New sheets added:
#   pca_params    eigenvectors, explained variance, component names
#   pca_train     PCA scores for all train rows
#   pca_test      PCA scores for all test rows
#
# PCA method:
#   - Input: kept sensors only (same as data_norm — drop std=0 sensors)
#   - Normalization: healthy-cycle z-score, per condition (same as v5)
#   - Fit PCA on train_normalized only
#   - Transform both train and test using train-fitted PCA
#   - Components kept: Option B — individual variance >= 5%
#     FD001: 2 components (82.9% variance)
#     FD002: 4 components (74.8% variance)
#     FD003: 3 components (87.8% variance)
#     FD004: 3 components (77.5% variance)
#
# Reference: Saxena et al., PHM08

import os
import shutil
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans


# ── constants ─────────────────────────────────────────────────────────────────

COLUMNS = [
  'unit_number', 'time_cycles',
  'op_setting_1', 'op_setting_2', 'op_setting_3',
  'T2', 'T24', 'T30', 'T50', 'P2', 'P15', 'P30',
  'Nf', 'Nc', 'epr', 'Ps30', 'phi',
  'NRf', 'NRc', 'BPR', 'farB', 'htBleed',
  'Nf_dmd', 'PCNfR_dmd', 'W31', 'W32',
]
ALL_SENSORS = COLUMNS[5:]
CAP         = 125
DATA_DIR    = '/home/claude/CMAPSSData'
OUT_DIR     = '/home/claude'

# Option B threshold
VARIANCE_THRESHOLD = 0.05

# styles
H_BLUE   = PatternFill('solid', start_color='2C4A7C', end_color='2C4A7C')
H_PURPLE = PatternFill('solid', start_color='6B1A6B', end_color='6B1A6B')
H_GREEN  = PatternFill('solid', start_color='1A6B3C', end_color='1A6B3C')
H_TEAL   = PatternFill('solid', start_color='0F6E56', end_color='0F6E56')
H_AMBER  = PatternFill('solid', start_color='854F0B', end_color='854F0B')
H_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
D_FONT   = Font(name='Arial', size=10)
THIN     = Side(style='thin', color='D3D1C7')
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── helpers ───────────────────────────────────────────────────────────────────

def hdr(ws, col, row, text, fill):
  c = ws.cell(row=row, column=col, value=text)
  c.fill = fill; c.font = H_FONT; c.border = BORDER
  c.alignment = Alignment(
    horizontal='center', vertical='center', wrap_text=True)
  return c


def dat(ws, row, col, value, fmt=None):
  c = ws.cell(row=row, column=col, value=value)
  c.font = D_FONT; c.border = BORDER
  c.alignment = Alignment(horizontal='right')
  if fmt:
    c.number_format = fmt
  return c


# ── data preparation ──────────────────────────────────────────────────────────

class CmapssLoader:

  def __init__(self, data_dir):
    self.data_dir = data_dir

  def calc_load(self, fd):
    train = pd.read_csv(
      os.path.join(self.data_dir, f'train_{fd}.txt'),
      sep=r'\s+', header=None, names=COLUMNS)
    test = pd.read_csv(
      os.path.join(self.data_dir, f'test_{fd}.txt'),
      sep=r'\s+', header=None, names=COLUMNS)
    return train, test


class CmapssPrep:

  def calc_prepare(self, train, test, n_conds):
    if n_conds > 1:
      ops = train[['op_setting_1', 'op_setting_2', 'op_setting_3']].values
      km  = KMeans(n_clusters=6, random_state=42, n_init=10)
      km.fit(ops)
      train['condition_id'] = km.predict(ops) + 1
      test['condition_id']  = km.predict(
        test[['op_setting_1', 'op_setting_2', 'op_setting_3']].values) + 1
    else:
      train['condition_id'] = 1
      test['condition_id']  = 1

    t_max = train.groupby('unit_number')['time_cycles'].transform('max')
    train['RUL_raw'] = (t_max - train['time_cycles']).astype(int)

    return train, test


class CmapssNorm:

  def calc_keep(self, train_df):
    return [s for s in ALL_SENSORS if train_df[s].std() >= 1e-6]

  def calc_stats(self, train_df, keep):
    healthy = train_df[train_df['RUL_raw'] >= CAP]
    stats   = {}
    for cond in sorted(train_df['condition_id'].unique()):
      subset = healthy[healthy['condition_id'] == cond]
      for s in keep:
        mu    = float(subset[s].mean())
        sigma = float(subset[s].std())
        if pd.isna(sigma) or sigma == 0:
          sigma = 1.0
        stats[(cond, s)] = (mu, sigma)
    return stats

  def calc_normalize(self, df, stats, keep):
    norm = df[keep].copy().astype(float)
    for cond in sorted(df['condition_id'].unique()):
      mask = (df['condition_id'] == cond).values
      idx  = df.index[mask]
      for s in keep:
        mu, sigma = stats[(cond, s)]
        norm.loc[idx, s] = ((df.loc[idx, s] - mu) / sigma).round(6)
    return norm.values


# ── PCA calculator ────────────────────────────────────────────────────────────

class PcaCalculator:

  def __init__(self):
    self.pca        = None
    self.n_keep     = None
    self.keep       = None
    self.var_ratio  = None
    self.cum_var    = None
    self.components = None  # component names: ['PC1', 'PC2', ...]

  def calc_fit(self, X_train, keep):
    self.keep  = keep
    self.pca   = PCA()
    self.pca.fit(X_train)

    self.var_ratio = self.pca.explained_variance_ratio_
    self.cum_var   = np.cumsum(self.var_ratio)

    # Option B: keep components where individual variance >= threshold
    self.n_keep = sum(1 for v in self.var_ratio
                      if v >= VARIANCE_THRESHOLD)
    if self.n_keep == 0:
      self.n_keep = 1

    self.components = [f'PC{i+1}' for i in range(self.n_keep)]

  def calc_transform(self, X):
    scores = self.pca.transform(X)[:, :self.n_keep]
    return np.round(scores, 6)

  def calc_loadings(self):
    # eigenvectors: shape (n_keep, n_sensors)
    return self.pca.components_[:self.n_keep]


# ── sheet builders ────────────────────────────────────────────────────────────

class PcaSheetBuilder:

  def build_pca_params_sheet(self, wb, pca_calc):
    ws = wb.create_sheet('pca_params')
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    n_comp    = pca_calc.n_keep
    n_sensors = len(pca_calc.keep)
    loadings  = pca_calc.calc_loadings()
    vr        = pca_calc.var_ratio
    cv        = pca_calc.cum_var

    # ── section 1: variance summary ───────────────────────────────
    ws.cell(row=1, column=1).value = 'Variance summary'
    ws.cell(row=1, column=1).font  = Font(
      bold=True, name='Arial', size=10, color='FFFFFF')
    ws.cell(row=1, column=1).fill  = H_BLUE
    ws.merge_cells(
      start_row=1, start_column=1, end_row=1, end_column=4)

    for col, label, fill in [
      (1, 'component', H_BLUE),
      (2, 'individual\nvariance', H_TEAL),
      (3, 'cumulative\nvariance', H_TEAL),
      (4, 'kept?', H_AMBER),
    ]:
      hdr(ws, col, 2, label, fill)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 8

    for i, v in enumerate(vr):
      row   = i + 3
      kept  = 'YES' if i < n_comp else 'no'
      fill  = H_GREEN if i < n_comp else None
      c = ws.cell(row=row, column=1, value=f'PC{i+1}')
      c.font = D_FONT; c.border = BORDER
      c.alignment = Alignment(horizontal='center')
      if i < n_comp:
        c.fill = H_GREEN
        c.font = H_FONT

      dat(ws, row, 2, round(float(v) * 100, 2), '0.00"%"')
      dat(ws, row, 3, round(float(cv[i]) * 100, 2), '0.00"%"')
      c4 = ws.cell(row=row, column=4, value=kept)
      c4.font = D_FONT; c4.border = BORDER
      c4.alignment = Alignment(horizontal='center')
      if i < n_comp:
        c4.fill = H_GREEN
        c4.font = H_FONT

    # ── section 2: loadings matrix ────────────────────────────────
    load_row_start = len(vr) + 5

    ws.cell(row=load_row_start, column=1).value = 'Eigenvectors (loadings)'
    ws.cell(row=load_row_start, column=1).font  = Font(
      bold=True, name='Arial', size=10, color='FFFFFF')
    ws.cell(row=load_row_start, column=1).fill  = H_PURPLE
    ws.merge_cells(
      start_row=load_row_start, start_column=1,
      end_row=load_row_start, end_column=n_comp + 1)

    # header: sensor | PC1 | PC2 | ...
    hdr(ws, 1, load_row_start + 1, 'sensor', H_PURPLE)
    for j in range(n_comp):
      hdr(ws, j + 2, load_row_start + 1, f'PC{j+1}', H_PURPLE)
      ws.column_dimensions[get_column_letter(j + 2)].width = 10

    # loadings data: one row per sensor
    for i, sensor in enumerate(pca_calc.keep):
      row = load_row_start + 2 + i
      c = ws.cell(row=row, column=1, value=sensor)
      c.font = D_FONT; c.border = BORDER
      c.alignment = Alignment(horizontal='left')
      for j in range(n_comp):
        dat(ws, row, j + 2,
            round(float(loadings[j, i]), 6), '0.000000')

    # color scale on loadings
    last_load_row = load_row_start + 1 + n_sensors
    last_col      = get_column_letter(n_comp + 1)
    ws.conditional_formatting.add(
      f'B{load_row_start + 2}:{last_col}{last_load_row}',
      ColorScaleRule(
        start_type='min', start_color='4F81BD',
        mid_type='num',   mid_value=0, mid_color='FFFFFF',
        end_type='max',   end_color='C0504D'))

    # notes
    note_row = last_load_row + 2
    notes = [
      f'PCA fitted on train_normalized ({n_sensors} sensors).',
      f'Option B: kept components with individual variance >= '
      f'{VARIANCE_THRESHOLD*100:.0f}%.',
      f'Kept {n_comp} of {len(vr)} components '
      f'({cv[n_comp-1]*100:.1f}% cumulative variance).',
      'Loadings: each column is one eigenvector.',
      'Positive loading = sensor increases → PC score increases.',
      'Negative loading = sensor increases → PC score decreases.',
    ]
    for i, note in enumerate(notes):
      c = ws.cell(row=note_row + i, column=1, value=note)
      c.font = Font(name='Arial', size=9, italic=True)
    ws.column_dimensions['A'].width = 55

    ws.freeze_panes = 'A3'

  def build_pca_scores_sheet(self, wb, sheet_name,
                              scores, meta_cols, meta_df,
                              pca_calc, source_label):
    ws = wb.create_sheet(sheet_name)
    ws.row_dimensions[1].height = 32

    # columns: meta columns + PC1..PCn
    all_cols  = meta_cols + pca_calc.components
    n_meta    = len(meta_cols)

    ws.append(all_cols)
    for col, h in enumerate(all_cols, 1):
      c = ws.cell(row=1, column=col)
      if col <= n_meta:
        c.fill = H_BLUE
      else:
        c.fill = H_GREEN
      c.font = H_FONT; c.border = BORDER
      c.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
      ws.column_dimensions[get_column_letter(col)].width = max(
        len(h) + 2, 10)

    ws.cell(row=1, column=n_meta + 1).comment = Comment(
      f'PCA scores — {source_label}.\n'
      f'Fitted on train_normalized only.\n'
      f'{pca_calc.n_keep} components kept (individual variance >= '
      f'{VARIANCE_THRESHOLD*100:.0f}%).\n'
      f'Cumulative variance explained: '
      f'{pca_calc.cum_var[pca_calc.n_keep-1]*100:.1f}%.\n'
      'Test scores use TRAIN-fitted PCA parameters.',
      'cmapss_dataset')
    ws.freeze_panes = 'C2'

    # write data rows
    meta_vals = meta_df[meta_cols].values.tolist()
    for i, (meta_row, score_row) in enumerate(
        zip(meta_vals, scores.tolist())):
      ws.append(meta_row + score_row)


# ── main builder ──────────────────────────────────────────────────────────────

class PcaBuilder:

  def __init__(self, loader, prep, norm_calc,
               pca_calc, sheet_builder, out_dir):
    self.loader        = loader
    self.prep          = prep
    self.norm_calc     = norm_calc
    self.pca_calc      = pca_calc
    self.sheet_builder = sheet_builder
    self.out_dir       = out_dir

  def build_file(self, fd, n_conds):
    src  = os.path.join(self.out_dir, f'cmapss_{fd}_v5.xlsx')
    dst  = os.path.join(self.out_dir, f'cmapss_{fd}_v6.xlsx')
    shutil.copy2(src, dst)

    train, test = self.loader.calc_load(fd)
    train, test = self.prep.calc_prepare(train, test, n_conds)

    keep       = self.norm_calc.calc_keep(train)
    stats      = self.norm_calc.calc_stats(train, keep)
    X_train    = self.norm_calc.calc_normalize(train, stats, keep)
    X_test     = self.norm_calc.calc_normalize(test,  stats, keep)

    # fit PCA on train, transform both
    self.pca_calc.calc_fit(X_train, keep)
    train_scores = self.pca_calc.calc_transform(X_train)
    test_scores  = self.pca_calc.calc_transform(X_test)

    # meta columns to carry alongside PCA scores
    meta_cols = ['unit_number', 'time_cycles', 'condition_id']

    wb = load_workbook(dst)

    # remove existing pca sheets if present
    for name in ['pca_params', 'pca_train', 'pca_test']:
      if name in wb.sheetnames:
        del wb[name]

    self.sheet_builder.build_pca_params_sheet(wb, self.pca_calc)
    self.sheet_builder.build_pca_scores_sheet(
      wb, 'pca_train', train_scores,
      meta_cols, train, self.pca_calc, 'train')
    self.sheet_builder.build_pca_scores_sheet(
      wb, 'pca_test', test_scores,
      meta_cols, test, self.pca_calc, 'test (train-fitted PCA)')

    # update about sheet
    self._update_about(wb, fd, self.pca_calc)

    wb.save(dst)
    size_kb = os.path.getsize(dst) // 1024
    return dst, size_kb, self.pca_calc.n_keep, \
           self.pca_calc.cum_var[self.pca_calc.n_keep - 1]

  def _update_about(self, wb, fd, pca_calc):
    if 'about' not in wb.sheetnames:
      return
    ws  = wb['about']
    row = ws.max_row + 2
    c   = ws.cell(row=row, column=1,
                  value=f'v6 additions: pca_params, pca_train, pca_test')
    c.font = Font(bold=True, name='Arial', size=10)
    notes = [
      f'  pca_params   eigenvectors and explained variance '
      f'({pca_calc.n_keep} components kept)',
      f'  pca_train    PCA scores for train rows',
      f'  pca_test     PCA scores for test rows (TRAIN-fitted PCA)',
      f'  Variance threshold: individual >= '
      f'{VARIANCE_THRESHOLD*100:.0f}% (Option B)',
      f'  Cumulative variance: '
      f'{pca_calc.cum_var[pca_calc.n_keep-1]*100:.1f}%',
    ]
    for i, note in enumerate(notes):
      ws.cell(row=row + 1 + i, column=1, value=note).font = Font(
        name='Arial', size=10)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
  loader        = CmapssLoader(DATA_DIR)
  prep          = CmapssPrep()
  norm_calc     = CmapssNorm()
  pca_calc      = PcaCalculator()
  sheet_builder = PcaSheetBuilder()
  builder       = PcaBuilder(
    loader, prep, norm_calc, pca_calc, sheet_builder, OUT_DIR)

  print(f"{'FD':6}  {'sensors':>7}  {'components':>10}  "
        f"{'cum_var':>8}  {'size KB':>8}  sheets")
  print('-' * 65)

  for fd, n_conds in [('FD001',1),('FD002',6),('FD003',1),('FD004',6)]:
    dst, size_kb, n_comp, cum_var = builder.build_file(fd, n_conds)
    wb    = load_workbook(dst, read_only=True)
    names = wb.sheetnames
    wb.close()
    print(
      f"{fd:6}  {len(norm_calc.calc_keep(CmapssLoader(DATA_DIR).calc_load(fd)[0])):>7}  "
      f"{n_comp:>10}  {cum_var*100:>7.1f}%  {size_kb:>8}  "
      f"{', '.join(names)}")

  print()
  print("Files written:")
  for fd in ['FD001', 'FD002', 'FD003', 'FD004']:
    p = os.path.join(OUT_DIR, f'cmapss_{fd}_v6.xlsx')
    print(f"  {p}")


raise SystemExit(main())
