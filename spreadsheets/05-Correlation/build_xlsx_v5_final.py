# build_xlsx_v5_final.py
#
# Builds cmapss_FDxxx_v5.xlsx for all four C-MAPSS sub-datasets.
#
# Sheets included (analytical only — raw data lives in HDF5):
#   train_normalized   z-scored sensor columns, all train rows
#   RUL                unit_number, RUL, last_cycle, total_life_est
#   stats              per-sensor per-condition raw and normalized stats
#   correlation        Pearson correlation matrix of normalized sensors
#   engine_summary     one row per train engine
#
# Raw train and test sheets are omitted — they are large and already
# stored in the HDF5 files. A note sheet documents this.
#
# Normalization: healthy-cycle-only z-score, per condition per sensor.
#   Matches build_xlsx_v5.py from the spreadsheet thread.

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.cluster import KMeans


# ── constants ─────────────────────────────────────────────────────────────────

COLUMNS = [
  'unit_number', 'time_cycles',
  'op_setting_1', 'op_setting_2', 'op_setting_3',
  'T2', 'T24', 'T30', 'T50',
  'P2', 'P15', 'P30',
  'Nf', 'Nc', 'epr', 'Ps30', 'phi',
  'NRf', 'NRc', 'BPR', 'farB', 'htBleed',
  'Nf_dmd', 'PCNfR_dmd', 'W31', 'W32',
]
SENSORS  = COLUMNS[5:]
CAP      = 125
DATA_DIR = '/home/claude/CMAPSSData'
OUT_DIR  = '/home/claude'

H_BLUE   = PatternFill('solid', start_color='2C4A7C', end_color='2C4A7C')
H_PURPLE = PatternFill('solid', start_color='6B1A6B', end_color='6B1A6B')
H_GREEN  = PatternFill('solid', start_color='1A6B3C', end_color='1A6B3C')
H_TEAL   = PatternFill('solid', start_color='0F6E56', end_color='0F6E56')
H_AMBER  = PatternFill('solid', start_color='854F0B', end_color='854F0B')
H_NOTE   = PatternFill('solid', start_color='FFF3CD', end_color='FFF3CD')
H_CALC   = PatternFill('solid', start_color='1A6B3C', end_color='1A6B3C')
H_FONT   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
N_FONT   = Font(bold=True, color='7B5800', name='Arial', size=10)
D_FONT   = Font(name='Arial', size=10)
THIN     = Side(style='thin', color='D3D1C7')
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── helpers ───────────────────────────────────────────────────────────────────

def hdr(ws, col, row, text, fill):
  c = ws.cell(row=row, column=col, value=text)
  c.fill = fill; c.font = H_FONT; c.border = BORDER
  c.alignment = Alignment(
    horizontal='center', vertical='center', wrap_text=True)
  return c


def dat(ws, row, col, value, fmt=None):
  c = ws.cell(row=row, column=col, value=value)
  c.font = D_FONT; c.border = BORDER
  c.alignment = Alignment(horizontal='right')
  if fmt:
    c.number_format = fmt
  return c


# ── data preparation ──────────────────────────────────────────────────────────

class CmapssLoader:

  def __init__(self, data_dir):
    self.data_dir = data_dir

  def calc_load(self, fd):
    train = pd.read_csv(
      os.path.join(self.data_dir, f'train_{fd}.txt'),
      sep=r'\s+', header=None, names=COLUMNS)
    test = pd.read_csv(
      os.path.join(self.data_dir, f'test_{fd}.txt'),
      sep=r'\s+', header=None, names=COLUMNS)
    rul = pd.read_csv(
      os.path.join(self.data_dir, f'RUL_{fd}.txt'),
      sep=r'\s+', header=None, names=['RUL'])
    rul.insert(0, 'unit_number', range(1, len(rul) + 1))
    return train, test, rul


class CmapssPrep:

  def calc_prepare(self, train, test, rul, n_conds):
    if n_conds > 1:
      ops = train[['op_setting_1', 'op_setting_2', 'op_setting_3']].values
      km  = KMeans(n_clusters=6, random_state=42, n_init=10)
      km.fit(ops)
      train['condition_id'] = km.predict(ops) + 1
      test['condition_id']  = km.predict(
        test[['op_setting_1', 'op_setting_2', 'op_setting_3']].values) + 1
    else:
      train['condition_id'] = 1
      test['condition_id']  = 1

    for df_ in [train, test]:
      df_['dT_compressor'] = (df_['T30'] - df_['T24']).round(4)
      df_['dT_turbine']    = (df_['T50'] - df_['T30']).round(4)

    t_max = train.groupby('unit_number')['time_cycles'].transform('max')
    train['RUL_raw']    = (t_max - train['time_cycles']).astype(int)
    train['RUL_capped'] = train['RUL_raw'].clip(upper=CAP).astype(int)
    train['cycle_pct']  = (train['time_cycles'] / t_max).round(4)

    last_cycle = test.groupby('unit_number')['time_cycles'].max()
    rul['last_cycle']     = rul['unit_number'].map(last_cycle).astype(int)
    rul['total_life_est'] = rul['last_cycle'] + rul['RUL']

    return train, test, rul


class CmapssNorm:

  def calc_stats(self, train_df):
    healthy = train_df[train_df['RUL_raw'] >= CAP]
    stats   = {}
    for cond in sorted(train_df['condition_id'].unique()):
      subset = healthy[healthy['condition_id'] == cond]
      for s in SENSORS:
        mu    = float(subset[s].mean())
        sigma = float(subset[s].std())
        if pd.isna(sigma) or sigma == 0:
          sigma = 1.0
        stats[(cond, s)] = (mu, sigma)
    return stats

  def calc_normalize(self, df, stats):
    df_norm = df.copy()
    for s in SENSORS:
      df_norm[s] = df_norm[s].astype(float)
    for cond in sorted(df['condition_id'].unique()):
      mask = df['condition_id'] == cond
      for s in SENSORS:
        mu, sigma = stats[(cond, s)]
        df_norm.loc[mask, s] = (
          (df.loc[mask, s] - mu) / sigma).round(6)
    return df_norm


# ── sheet builders ────────────────────────────────────────────────────────────

class SheetBuilder:

  def build_note_sheet(self, wb, fd):
    ws = wb.create_sheet('about', 0)
    ws.column_dimensions['A'].width = 80
    ws.row_dimensions[1].height = 24

    notes = [
      [f'C-MAPSS Dataset — {fd} — v5'],
      [''],
      ['Sheets in this workbook:'],
      ['  about              this sheet'],
      ['  train_normalized   z-scored sensor columns, all train rows'],
      ['  test_normalized    z-scored sensor columns, all test rows (TRAIN params)'],
      ['  RUL                unit_number, RUL, last_cycle, total_life_est'],
      ['  stats              per-sensor, per-condition stats (raw + normalized)'],
      ['  correlation        Pearson correlation matrix'],
      ['  engine_summary     one row per train engine'],
      [''],
      ['Normalization: z-score on healthy cycles (RUL_raw >= 125), per condition.'],
      ['Test set normalized using TRAIN parameters — not fitted on test data.'],
      ['Raw train and test data omitted — stored in cmapss_norm_FDxxx.h5'],
      ['Reference: Saxena et al., PHM08 — C-MAPSS dataset.'],
    ]
    for i, row in enumerate(notes, 1):
      c = ws.cell(row=i, column=1, value=row[0])
      if i == 1:
        c.font = Font(bold=True, name='Arial', size=12)
      else:
        c.font = Font(name='Arial', size=10)

  def build_norm_sheet(self, wb, train_norm):
    ws = wb.create_sheet('train_normalized')
    hdrs = list(train_norm.columns)
    ws.row_dimensions[1].height = 32

    ws.append(hdrs)
    for col, h in enumerate(hdrs, 1):
      c = ws.cell(row=1, column=col)
      if col > 26:
        c.fill = H_CALC
      elif col > 5:
        c.fill = H_PURPLE
      else:
        c.fill = H_BLUE
      c.font = H_FONT
      c.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
      ws.column_dimensions[get_column_letter(col)].width = max(
        len(h) + 2, 10)

    ws.cell(row=1, column=6).comment = Comment(
      'Sensors normalized: z = (x - mu_c) / sigma_c\n'
      'mu_c, sigma_c from healthy cycles (RUL_raw >= 125) per condition.\n'
      'Cols 1-5 unchanged.',
      'cmapss_dataset')
    ws.freeze_panes = 'C2'

    for row in train_norm.values.tolist():
      ws.append(row)

  def build_rul_sheet(self, wb, rul):
    ws   = wb.create_sheet('RUL')
    hdrs = list(rul.columns)
    ws.row_dimensions[1].height = 28
    ws.append(hdrs)

    for col, h in enumerate(hdrs, 1):
      c = ws.cell(row=1, column=col)
      c.fill = H_BLUE if col <= 2 else H_CALC
      c.font = H_FONT
      c.border = BORDER
      c.alignment = Alignment(horizontal='center', vertical='center')
      ws.column_dimensions[get_column_letter(col)].width = max(
        len(h) + 3, 12)

    ws.cell(row=1, column=3).comment = Comment(
      'Derived from test sheet: max(time_cycles) per engine.',
      'cmapss_dataset')
    ws.cell(row=1, column=4).comment = Comment(
      '= last_cycle + RUL\nEstimated full engine life.',
      'cmapss_dataset')

    for row in rul.values.tolist():
      ws.append(row)
    ws.freeze_panes = 'A2'

  def build_stats_sheet(self, wb, train, train_norm):
    ws = wb.create_sheet('stats')
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 28

    for rng, label, fill in [
      ('A1:A2', 'Sensor',     H_BLUE),
      ('B1:B2', 'Condition',  H_BLUE),
      ('C1:F1', 'Raw values', H_TEAL),
      ('G1:J1', 'Normalized', H_PURPLE),
    ]:
      ws.merge_cells(rng)
      c = ws[rng.split(':')[0]]
      c.value = label; c.fill = fill; c.font = H_FONT
      c.alignment = Alignment(horizontal='center', vertical='center')

    for col, label, fill in [
      (3,'min',H_TEAL), (4,'max',H_TEAL),
      (5,'mean',H_TEAL),(6,'std',H_TEAL),
      (7,'min',H_PURPLE),(8,'max',H_PURPLE),
      (9,'mean',H_PURPLE),(10,'std',H_PURPLE),
    ]:
      hdr(ws, col, 2, label, fill)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    for c in range(3, 11):
      ws.column_dimensions[get_column_letter(c)].width = 10

    row = 3
    for sensor in SENSORS:
      for cond in sorted(train['condition_id'].unique()):
        mask   = train['condition_id'] == cond
        mask_n = train_norm['condition_id'] == cond
        rv     = train.loc[mask, sensor]
        nv     = train_norm.loc[mask_n, sensor]

        c = ws.cell(row=row, column=1, value=sensor)
        c.font = D_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal='left')
        c = ws.cell(row=row, column=2, value=int(cond))
        c.font = D_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal='center')

        for col, val in [
          (3, rv.min()),  (4, rv.max()),
          (5, rv.mean()), (6, rv.std()),
          (7, nv.min()),  (8, nv.max()),
          (9, nv.mean()), (10, nv.std()),
        ]:
          dat(ws, row, col, round(float(val), 4), '0.0000')
        row += 1

    ws.freeze_panes = 'C3'
    ws.conditional_formatting.add(
      f'I3:I{row - 1}',
      ColorScaleRule(
        start_type='min', start_color='4F81BD',
        mid_type='num',   mid_value=0, mid_color='FFFFFF',
        end_type='max',   end_color='C0504D'))

  def build_correlation_sheet(self, wb, train_norm):
    ws   = wb.create_sheet('correlation')
    corr = train_norm[SENSORS].corr().round(4)

    ws.row_dimensions[1].height = 80
    ws.column_dimensions['A'].width = 14

    c = ws.cell(row=1, column=1, value='sensor')
    c.fill = H_BLUE; c.font = H_FONT; c.border = BORDER
    c.alignment = Alignment(horizontal='center', vertical='center')

    for i, s in enumerate(SENSORS, start=2):
      c = ws.cell(row=1, column=i, value=s)
      c.fill = H_PURPLE; c.font = H_FONT; c.border = BORDER
      c.alignment = Alignment(
        horizontal='center', vertical='center', text_rotation=90)
      ws.column_dimensions[get_column_letter(i)].width = 6

      c2 = ws.cell(row=i, column=1, value=s)
      c2.fill = H_PURPLE; c2.font = H_FONT; c2.border = BORDER
      c2.alignment = Alignment(horizontal='left', vertical='center')

    for r, s1 in enumerate(SENSORS, start=2):
      for ci, s2 in enumerate(SENSORS, start=2):
        val  = corr.loc[s1, s2]
        cell = ws.cell(row=r, column=ci, value=float(val))
        cell.font = Font(name='Arial', size=9)
        cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    ws.freeze_panes = 'B2'
    last = len(SENSORS) + 1
    ws.conditional_formatting.add(
      f'B2:{get_column_letter(last)}{last}',
      ColorScaleRule(
        start_type='num', start_value=-1, start_color='4F81BD',
        mid_type='num',   mid_value=0,    mid_color='FFFFFF',
        end_type='num',   end_value=1,    end_color='C0504D'))

    ws.cell(row=1, column=1).comment = Comment(
      'Pearson correlation of normalized sensor values.\n'
      'Red = strong positive (+1).\n'
      'Blue = strong negative (-1).\n'
      'White = no correlation (0).',
      'cmapss_dataset')

  def build_test_norm_sheet(self, wb, test_norm):
    ws = wb.create_sheet('test_normalized')
    hdrs = list(test_norm.columns)
    ws.row_dimensions[1].height = 32

    ws.append(hdrs)
    for col, h in enumerate(hdrs, 1):
      c = ws.cell(row=1, column=col)
      if col > 26:
        c.fill = H_CALC
      elif col > 5:
        c.fill = H_PURPLE
      else:
        c.fill = H_BLUE
      c.font = H_FONT
      c.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
      ws.column_dimensions[get_column_letter(col)].width = max(
        len(h) + 2, 10)

    ws.cell(row=1, column=6).comment = Comment(
      'Test sensors normalized using TRAIN parameters.\n'
      'z = (x - mu_c) / sigma_c\n'
      'mu_c, sigma_c from train healthy cycles (RUL_raw >= 125) per condition.\n'
      'NOT fitted on test data — train params applied to test.\n'
      'Cols 1-5 unchanged.',
      'cmapss_dataset')
    ws.freeze_panes = 'C2'

    for row in test_norm.values.tolist():
      ws.append(row)

  def build_engine_summary_sheet(self, wb, train, train_norm, rul):
    ws = wb.create_sheet('engine_summary')
    ws.row_dimensions[1].height = 32

    hdrs = [
      ('unit_number',   H_BLUE),
      ('total_life',    H_TEAL),
      ('n_conditions',  H_TEAL),
      ('dominant_cond', H_TEAL),
      ('T30_mean_raw',  H_GREEN),
      ('T30_mean_norm', H_PURPLE),
      ('phi_mean_raw',  H_GREEN),
      ('phi_mean_norm', H_PURPLE),
      ('RUL_at_50pct',  H_AMBER),
      ('RUL_at_75pct',  H_AMBER),
      ('RUL_test',      H_AMBER),
    ]
    for col, (label, fill) in enumerate(hdrs, 1):
      hdr(ws, col, 1, label, fill)
      ws.column_dimensions[get_column_letter(col)].width = max(
        len(label) + 3, 12)

    for row_idx, unit in enumerate(
        sorted(train['unit_number'].unique()), start=2):
      ut  = train[train['unit_number'] == unit]
      utn = train_norm[train_norm['unit_number'] == unit]

      total_life = int(ut['time_cycles'].max())
      n_cond     = int(ut['condition_id'].nunique())
      dom_cond   = int(ut['condition_id'].mode().iloc[0])
      t30_raw    = round(float(ut['T30'].mean()), 4)
      t30_norm   = round(float(utn['T30'].mean()), 4)
      phi_raw    = round(float(ut['phi'].mean()), 4)
      phi_norm   = round(float(utn['phi'].mean()), 4)

      t50  = total_life * 0.50
      t75  = total_life * 0.75
      r50  = ut[ut['time_cycles'] <= t50]['RUL_raw'].min()
      r75  = ut[ut['time_cycles'] <= t75]['RUL_raw'].min()
      r50  = int(r50) if not pd.isna(r50) else ''
      r75  = int(r75) if not pd.isna(r75) else ''

      rul_val = ''
      if unit in rul['unit_number'].values:
        rul_val = int(
          rul[rul['unit_number'] == unit]['RUL'].values[0])

      vals = [unit, total_life, n_cond, dom_cond,
              t30_raw, t30_norm, phi_raw, phi_norm,
              r50, r75, rul_val]
      for col, val in enumerate(vals, 1):
        dat(ws, row_idx, col, val)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(hdrs))}1'
    ws.cell(row=1, column=1).comment = Comment(
      'One row per training engine.\n'
      'RUL_at_50pct/75pct: remaining life at 50%/75% of total life.\n'
      'RUL_test: from RUL file (blank — train/test are different instances).',
      'cmapss_dataset')


# ── main ──────────────────────────────────────────────────────────────────────

def main():
  loader  = CmapssLoader(DATA_DIR)
  prep    = CmapssPrep()
  norm    = CmapssNorm()
  builder = SheetBuilder()

  FD_LIST = [('FD001', 1), ('FD002', 6), ('FD003', 1), ('FD004', 6)]

  for fd, n_conds in FD_LIST:
    print(f"{fd}...", flush=True)

    train, test, rul = loader.calc_load(fd)
    train, test, rul = prep.calc_prepare(train, test, rul, n_conds)
    stats            = norm.calc_stats(train)
    train_norm       = norm.calc_normalize(train, stats)
    test_norm        = norm.calc_normalize(test,  stats)

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    builder.build_note_sheet(wb, fd)
    builder.build_norm_sheet(wb, train_norm)
    builder.build_test_norm_sheet(wb, test_norm)
    builder.build_rul_sheet(wb, rul)
    builder.build_stats_sheet(wb, train, train_norm)
    builder.build_correlation_sheet(wb, train_norm)
    builder.build_engine_summary_sheet(wb, train, train_norm, rul)

    path = os.path.join(OUT_DIR, f'cmapss_{fd}_v5.xlsx')
    wb.save(path)
    size_kb = os.path.getsize(path) // 1024
    print(f"  saved {path}  ({size_kb} KB)")

  print("\nAll done.")


raise SystemExit(main())
