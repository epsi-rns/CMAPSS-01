import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, column_index_from_string
from sklearn.cluster import KMeans
import shutil

COLUMNS = [
    'unit_number','time_cycles',
    'op_setting_1','op_setting_2','op_setting_3',
    'T2','T24','T30','T50',
    'P2','P15','P30',
    'Nf','Nc','epr','Ps30','phi',
    'NRf','NRc','BPR','farB','htBleed',
    'Nf_dmd','PCNfR_dmd','W31','W32'
]
SENSORS = COLUMNS[5:]   # cols 6-26, indices 5-25
CAP     = 125

H_FILL  = PatternFill('solid', start_color='2C4A7C', end_color='2C4A7C')
H_NORM  = PatternFill('solid', start_color='6B1A6B', end_color='6B1A6B')
H_CALC  = PatternFill('solid', start_color='1A6B3C', end_color='1A6B3C')
H_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)


def compute_norm_stats(train_df, sensor_cols, cond_col='condition_id',
                       rul_col='RUL_raw'):
    """
    Compute per-condition mean and std from healthy cycles only.
    Healthy = RUL_raw >= 125.
    Returns dict: {(condition, sensor): (mu, sigma)}
    """
    healthy = train_df[train_df[rul_col] >= CAP]
    stats = {}
    for cond in sorted(train_df[cond_col].unique()):
        subset = healthy[healthy[cond_col] == cond]
        for s in sensor_cols:
            mu    = subset[s].mean()
            sigma = subset[s].std()
            if pd.isna(sigma) or sigma == 0:
                sigma = 1.0   # avoid division by zero
            stats[(cond, s)] = (mu, sigma)
    return stats


def normalize_df(df, stats, sensor_cols, cond_col='condition_id'):
    """Apply condition-wise z-score normalization to sensor columns."""
    df_norm = df.copy()
    for s in sensor_cols:
        df_norm[s] = df.apply(
            lambda row: (row[s] - stats[(row[cond_col], s)][0])
                        / stats[(row[cond_col], s)][1],
            axis=1
        ).round(6)
    return df_norm


def write_norm_header(ws, headers, base_ncols):
    ws.row_dimensions[1].height = 32
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        if col > base_ncols:
            c.fill = H_CALC
        elif col > 5:
            c.fill = H_NORM   # purple = normalized sensors
        else:
            c.fill = H_FILL   # blue = identifiers/op settings
        c.font = H_FONT
        c.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(
            len(h) + 2, 10)
    # comment on first sensor header
    ws.cell(row=1, column=6).comment = Comment(
        "Sensors normalized: z = (x - mu_c) / sigma_c\n"
        "mu_c, sigma_c computed from healthy cycles (RUL_raw >= 125)\n"
        "per condition cluster (condition_id).\n"
        "Cols 1-5 (identifiers + op settings) unchanged.",
        "cmapss_dataset")
    ws.freeze_panes = 'C2'


# ------------------------------------------------------------------
for fd, n_conds in [('FD001',1),('FD002',6),('FD003',1),('FD004',6)]:
    print(f"{fd}...", flush=True)

    # --- load raw data ---
    train = pd.read_csv(
        f'/tmp/train_{fd}.txt', sep=r'\s+', header=None, names=COLUMNS)
    test  = pd.read_csv(
        f'/tmp/test_{fd}.txt',  sep=r'\s+', header=None, names=COLUMNS)
    rul   = pd.read_csv(
        f'/tmp/RUL_{fd}.txt',   sep=r'\s+', header=None, names=['RUL'])
    rul.insert(0, 'unit_number', range(1, len(rul)+1))

    # --- condition_id ---
    if n_conds > 1:
        ops = train[['op_setting_1','op_setting_2','op_setting_3']].values
        km  = KMeans(n_clusters=6, random_state=42, n_init=10)
        km.fit(ops)
        train['condition_id'] = km.predict(ops) + 1
        test['condition_id']  = km.predict(
            test[['op_setting_1','op_setting_2','op_setting_3']].values) + 1
    else:
        train['condition_id'] = 1
        test['condition_id']  = 1

    # --- derived sensor features ---
    for df_ in [train, test]:
        df_['dT_compressor'] = (df_['T30'] - df_['T24']).round(4)
        df_['dT_turbine']    = (df_['T50'] - df_['T30']).round(4)

    # --- RUL columns (train only) ---
    t_max = train.groupby('unit_number')['time_cycles'].transform('max')
    train['RUL_raw']    = (t_max - train['time_cycles']).astype(int)
    train['RUL_capped'] = train['RUL_raw'].clip(upper=CAP).astype(int)
    train['cycle_pct']  = (train['time_cycles'] / t_max).round(4)

    # --- RUL sheet extra columns ---
    last_cycle = test.groupby('unit_number')['time_cycles'].max()
    rul['last_cycle']     = rul['unit_number'].map(last_cycle).astype(int)
    rul['total_life_est'] = rul['last_cycle'] + rul['RUL']

    # --- normalization stats (from train healthy cycles) ---
    stats = compute_norm_stats(train, SENSORS)

    # --- apply normalization ---
    train_norm = normalize_df(train, stats, SENSORS)

    # --- start from v3 as base (copy it) ---
    src  = f'/home/claude/cmapss_{fd}_v3.xlsx'
    dest = f'/home/claude/cmapss_{fd}_v4.xlsx'
    shutil.copy(src, dest)

    # --- open and add normalized sheet ---
    wb = load_workbook(dest)

    ws_n = wb.create_sheet('train_normalized')
    headers_n = list(train_norm.columns)
    write_norm_header(ws_n, headers_n, 26)
    for row in train_norm.itertuples(index=False):
        ws_n.append(list(row))

    wb.save(dest)
    print(f"  saved {dest} — norm rows: {len(train_norm)}", flush=True)

print("All done.")
