import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from sklearn.cluster import KMeans

COLUMNS = [
    'unit_number','time_cycles',
    'op_setting_1','op_setting_2','op_setting_3',
    'T2','T24','T30','T50',
    'P2','P15','P30',
    'Nf','Nc','epr','Ps30','phi',
    'NRf','NRc','BPR','farB','htBleed',
    'Nf_dmd','PCNfR_dmd','W31','W32'
]
CAP = 125

H_FILL = PatternFill('solid', start_color='2C4A7C', end_color='2C4A7C')
H_CALC = PatternFill('solid', start_color='1A6B3C', end_color='1A6B3C')
H_NOTE = PatternFill('solid', start_color='FFF3CD', end_color='FFF3CD')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
N_FONT = Font(bold=True, color='7B5800', name='Arial', size=10)


def style_header(ws, headers, base_ncols):
    ws.row_dimensions[1].height = 32
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        if h == 'condition_id':
            c.fill = H_NOTE
            c.font = N_FONT
            c.comment = Comment(
                "KMeans(k=6) on op_setting_1/2/3.\n"
                "Fitted on train, applied to test.\n"
                "Labels 1-6 = six C-MAPSS flight conditions.\n"
                "Not reproducible as a spreadsheet formula.",
                "cmapss_dataset")
        elif col > base_ncols:
            c.fill = H_CALC
            c.font = H_FONT
        else:
            c.fill = H_FILL
            c.font = H_FONT
        c.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = max(len(h)+2, 10)


def write_sheet(ws, df, base_ncols, freeze='C2'):
    headers = list(df.columns)
    ws.append(headers)
    style_header(ws, headers, base_ncols)
    for row in df.itertuples(index=False):
        ws.append(list(row))
    ws.freeze_panes = freeze


for fd, n_conds in [('FD001',1),('FD002',6),('FD003',1),('FD004',6)]:
    print(f"{fd}...", flush=True)

    train = pd.read_csv(
        f'/tmp/train_{fd}.txt', sep=r'\s+', header=None, names=COLUMNS)
    test  = pd.read_csv(
        f'/tmp/test_{fd}.txt',  sep=r'\s+', header=None, names=COLUMNS)
    rul   = pd.read_csv(
        f'/tmp/RUL_{fd}.txt',   sep=r'\s+', header=None, names=['RUL'])
    rul.insert(0, 'unit_number', range(1, len(rul)+1))

    # --- condition_id ---
    if n_conds > 1:
        ops = train[['op_setting_1','op_setting_2','op_setting_3']].values
        km  = KMeans(n_clusters=6, random_state=42, n_init=10)
        km.fit(ops)
        train['condition_id'] = km.predict(ops) + 1
        test['condition_id']  = km.predict(
            test[['op_setting_1','op_setting_2','op_setting_3']].values) + 1
    else:
        train['condition_id'] = 1
        test['condition_id']  = 1

    # --- derived sensor features ---
    for df_ in [train, test]:
        df_['dT_compressor'] = (df_['T30'] - df_['T24']).round(4)
        df_['dT_turbine']    = (df_['T50'] - df_['T30']).round(4)

    # --- RUL columns (train only) ---
    t_max = train.groupby('unit_number')['time_cycles'].transform('max')
    train['RUL_raw']    = (t_max - train['time_cycles']).astype(int)
    train['RUL_capped'] = train['RUL_raw'].clip(upper=CAP).astype(int)
    train['cycle_pct']  = (train['time_cycles'] / t_max).round(4)

    # --- RUL sheet extra columns ---
    last_cycle = test.groupby('unit_number')['time_cycles'].max()
    rul['last_cycle']      = rul['unit_number'].map(last_cycle).astype(int)
    rul['total_life_est']  = rul['last_cycle'] + rul['RUL']

    # --- build workbook ---
    wb = Workbook()

    ws_tr = wb.active; ws_tr.title = 'train'
    write_sheet(ws_tr, train, 26, freeze='C2')

    ws_te = wb.create_sheet('test')
    write_sheet(ws_te, test, 26, freeze='C2')

    ws_r = wb.create_sheet('RUL')
    # RUL sheet: cols 1-2 original (blue), cols 3-4 derived (green)
    ws_r.append(['unit_number','RUL','last_cycle','total_life_est'])
    style_header(ws_r, ['unit_number','RUL','last_cycle','total_life_est'], 2)
    ws_r.cell(row=1, column=3).comment = Comment(
        "Derived from test sheet: max(time_cycles) per engine.\n"
        "The cycle at which the test trajectory was cut.\n"
        "Not stated in paper — logical consequence of problem definition.",
        "cmapss_dataset")
    ws_r.cell(row=1, column=4).comment = Comment(
        "= last_cycle + RUL\n"
        "Estimated full engine life (observed + remaining).\n"
        "Not stated in paper — derived for interpretability.",
        "cmapss_dataset")
    for row in rul.itertuples(index=False):
        ws_r.append(list(row))
    ws_r.freeze_panes = 'A2'

    path = f'/home/claude/cmapss_{fd}_v3.xlsx'
    wb.save(path)
    print(f"  saved {path} — train:{len(train)} test:{len(test)} RUL:{len(rul)}",
          flush=True)

print("All done.")
